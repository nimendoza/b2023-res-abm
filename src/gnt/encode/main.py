from src.gnt.encode.constants import *

from src.cls.main       import GradeLevel
from src.cls.main       import Subject
from src.cls.main       import Category
from src.cls.main       import Student
from src.cls.main       import Group
from src.cls.main       import Shift
from src.cls.main       import Section
from src.cls.main       import ParallelSession
from src.cls.main       import Capacity
from openpyxl           import load_workbook
from src.gnt.write.main import writer_agent


class EncodeAgent:
    def __init__(self) -> None:
        self._grade_levels = dict[str, GradeLevel]()
        self._subjects     = dict[str, list[set[tuple[GradeLevel, bool, str]] | Subject]]()
        self._categories   = dict[str, list[set[tuple[GradeLevel, str]] | Category]]()
        self._students     = dict[GradeLevel, list[Student]]()
        self._groups       = dict[GradeLevel, dict[Subject | Category, dict[str, Group]]]()
        self._shifts       = dict[str, Shift]()

    def __sort__(self) -> None:
        ...

    def __reset__(self) -> None:
        ...

    def __reset_students__(self) -> None:
        ...

    @property
    def grade_levels(self) -> dict[str, GradeLevel]:
        return self._grade_levels

    @property
    def subjects(self) -> dict[str, list[set[tuple[GradeLevel, bool, str]] | Subject]]:
        return self._subjects

    @property
    def categories(self) -> dict[str, list[set[tuple[GradeLevel, str]] | Category]]:
        return self._categories

    @property
    def students(self) -> dict[GradeLevel, list[Student]]:
        return self._students

    @property
    def groups(self) -> dict[GradeLevel, dict[Subject | Category, dict[str, Group]]]:
        return self._groups

    @property
    def shifts(self) -> dict[str, Shift]:
        return self._shifts

    def xlsx(
        self,
        path : str,
        sheet: str
    ) -> list[list[str | None]]:
        workbook  = load_workbook(filename=path, data_only=True)
        worksheet = workbook[sheet]
        data      = list()
        for row in worksheet.rows:
            line = list()
            for cell in row:
                if cell.value is None:
                    line.append(cell.value)
                elif isinstance(cell.value, float):
                    line.append(str(int(cell.value)))
                else:
                    line.append(str(cell.value))
            if not all(
                cell == 'None'
                    for cell in line
            ):
                data.append(line)
        workbook.close()
        return data

    def find(self, name: str) -> Subject | Category:
        if name in self.categories:
            return self.categories[name][1]
        else:
            if name in self.subjects:
                return self.subjects[name][1]
            else:
                for data, subject in self.subjects.values():
                    if str(subject) == name:
                        return subject

    def encode_subjects(
        self,
        path : str,
        reset: bool = None
    ) -> None:
        if reset:
            self.__reset__()
        
        for sheet in writer_agent.get_sheetnames(path):
            if sheet.find(GRADE_LEVEL) != -1:
                data        = self.xlsx(path, sheet)
                grade_level = int(data[0][1])
                self.grade_levels[str(grade_level)] = GradeLevel(
                    grade_level=grade_level,
                    to_rank=set(data[7][c] for c in range(3, 3 + int(data[7][2]))),
                    category_types=set(data[4][c] for c in range(3, 3 + int(data[4][2]))),
                    subject_types=set(data[3][c] for c in range(3, 3 + int(data[3][2])))
                )
        
        data = self.xlsx(path, SHIFTS)
        for r in range(1, len(data)):
            name  = data[r][0]
            shift = Shift(name)
            for c in range(2, 2 + int(data[r][1])):
                shift.add(data[r][c])
            self.shifts[name] = shift

        data = self.xlsx(path, CLASSIFICATION)
        for r in range(2, len(data)):
            name  = data[r][0]
            level = None if not data[r][1].isdigit() else int(data[r][1])
            max_group_members = None if not data[r][2].isdigit() else int(data[r][2])
            
            c = 3
            while data[1][c] == CATEGORY:
                grade_level = self.grade_levels[data[1][c + 1]]
                type        = data[1][c + 2]
                if any(cell.upper() == YES for cell in data[r][c:c + 3] if cell):
                    if name not in self.categories:
                        category = Category(
                            name=name,
                            types={type},
                            teaches={grade_level}
                        )
                        self.categories[name] = [{(grade_level, type)}, category]
                    else:
                        self.categories[name][0].add((grade_level, type))
                        self.categories[name][1].teaches.add(grade_level)
                        self.categories[name][1].types.add(type)
                    grade_level.category_types.add(type)
                c += 3
            while c < len(data[1]) and data[1][c] == SUBJECT:
                grade_level = self.grade_levels[data[1][c + 1]]
                ranked = data[1][c + 2].upper() == YES
                type = data[1][c + 3]
                if any(cell.upper() == YES for cell in data[r][c:c + 4] if cell):
                    if name not in self.subjects:
                        subject = Subject(
                            name=name,
                            types={type},
                            level=level,
                            max_group_members=max_group_members,
                            teaches={grade_level}
                        )
                        self.subjects[name] = [{(grade_level, ranked, type)}, subject]
                    else:
                        subject_ = Subject(name=name, level=level, types={})
                        if str(self.subjects[name][1]) == str(subject_):
                            self.subjects[name][0].add((grade_level, ranked, type))
                            self.subjects[name][1].teaches.add(grade_level)
                            self.subjects[name][1].types.add(type)
                        else:
                            name_ = str(subject_)
                            if name_ not in self.subjects:
                                subject_ = Subject(
                                    name=name,
                                    types={type},
                                    level=level,
                                    max_group_members=max_group_members,
                                    teaches={grade_level}
                                )
                                self.subjects[str(subject_)] = [
                                    {(grade_level, ranked, type)},
                                    subject_
                                ]
                            else:
                                self.subjects[name_][0].add((grade_level, ranked, type))
                                self.subjects[name_][1].teaches.add(grade_level)
                                self.subjects[name_][1].types.add(type)
                    grade_level.subject_types.add(type)
                    if ranked:
                        grade_level.to_rank.add(type)
                c += 4

        data = self.xlsx(path, PREREQUISITES)
        for r in range(1, len(data)):
            object = self.find(data[r][0])
            for c in range(2, 2 + int(data[r][1])):
                prerequisite = tuple(self.find(name) for name in data[r][c].split(DELIMITER))
                object.add_prerequisites(prerequisite)

        data = self.xlsx(path, NOT_ALONGSIDE)
        for r in range(1, len(data)):
            object = self.find(data[r][0])
            for c in range(2, 2 + int(data[r][1])):
                not_alongside = self.find(data[r][c])
                object.add_not_alongside(not_alongside)

        data = self.xlsx(path, SECTIONS)
        for r in range(2, len(data)):
            object = self.find(data[r][0])
            for c in range(2, 2 + int(data[r][1]), 6):
                shift = self.shifts[data[r][c + 2]]
                index = None if not data[r][c + 1].isdigit() else int(data[r][c + 1])
                object.add_section(Section(
                    shift=shift,
                    parallel_session=ParallelSession(
                        shift=shift,
                        partition=data[r][c],
                        index=index
                    ),
                    capacity=Capacity(
                        minimum=int(data[r][c + 3]),
                        ideal=int(data[r][c + 4]),
                        maximum=int(data[r][c + 5])
                    ),
                    parent=object
                ))

    def encode_students(
        self,
        path : str,
        reset: bool = None
    ) -> None:
        if reset:
            self.__reset_students__()

        # Groups need to be processed first
        for sheet in writer_agent.get_sheetnames(path):
            if sheet.find(GROUPS) != -1:
                data   = self.xlsx(path, sheet)
                parent = self.find(data[0][1])
                for grade_level in parent.teaches:
                    if grade_level not in self.groups:
                        self.groups[grade_level] = dict[Subject | Category, dict[str, Group]]()
                if parent not in self.groups[grade_level]:
                    self.groups[grade_level][parent] = dict[str, Group]()
                for r in range(2, len(data)):
                    id = data[r][0]
                    group = Group(
                        id=id,
                        parent=parent,
                        required=data[r][1].upper() == YES
                    )
                    self.groups[grade_level][parent][id] = group
        for sheet in writer_agent.get_sheetnames(path):
            if sheet.find(GRADE_LEVEL) != -1:
                data = self.xlsx(path, sheet)
                for r in range(2, len(data)):
                    grade_level = self.grade_levels[data[r][0]]
                    if grade_level not in self.students:
                        self.students[grade_level] = list[Student]()

                    id      = data[r][1]
                    student = Student(id=id, grade_level=grade_level)

                    c = 2
                    while data[0][c].upper() == GROUP.upper():
                        parent = self.find(data[1][c])
                        id_    = data[r][c]
                        if id_ in self.groups[grade_level][parent]:
                            self.groups[grade_level][parent][id_].add_student(student)
                        c += 1
                    while data[0][c].upper() == PREVIOUS_YEAR.upper():
                        if student.previous is None:
                            student_ = Student(
                                id=id, 
                                grade_level=self.grade_levels[str(grade_level.grade_level - 1)]
                            )
                            student.previous = student_
                        student.previous.attends.add(self.find(data[r][c]))
                        c += 1
                    for c_ in range(c, len(data[r])):
                        type = data[1][c_]
                        if data[r][c_] is None:
                            continue
                        object = self.find(data[r][c_])
                        if type in grade_level.to_rank:
                            student.rankings.add(type, object)
                        else:
                            if isinstance(object, Subject):
                                student.add_subject(type, object)
                            else:
                                student.add_category(type, object)
                    self.students[grade_level].append(student)


encode_agent = EncodeAgent()