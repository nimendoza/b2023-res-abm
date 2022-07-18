from src.gnt.write.constants import *

from xlsxwriter.workbook import Workbook
from typing              import Any
from os.path             import exists
from openpyxl            import load_workbook
from openpyxl.utils      import get_column_letter
from os                  import walk

class WriterAgent:
    def as_text(self, value: Any) -> str:
        if value is None:
            return ''
        else:
            return str(value)

    def generate_xlsx(self, path: str) -> None:
        workbook = Workbook(path)
        workbook.close()

    def write_to_xlsx(
        self,
        path: str,
        sheet: str,
        data: list[list[Any]]
    ) -> None:
        if not exists(path):
            self.generate_xlsx(path)

        workbook = load_workbook(filepath=path)
        if DEFAULT_SHEETNAME in workbook.sheetnames:
            workbook.remove(workbook[DEFAULT_SHEETNAME])

        worksheet = workbook.create_sheet(title=sheet)
        for row in data:
            worksheet.append(
                self.as_text(cell)
                    for cell in row
            )
        for column_cells in worksheet.columns:
            worksheet.column_dimensions[
                get_column_letter(column_cells[0].column)
            ].width = max(
                len(repr(cell.value))
                    for cell in column_cells
            )

        workbook.save(path)
        workbook.close()

    def find_filename(
        self,
        directory:     str,
        name_template: str
    ) -> str:
        path_template = '{}\\{}'.format(directory, name_template)
        for index in range(len(next(walk(directory))[2]) + 1):
            if not exists(path_template.format(index)):
                return path_template.format(index)

    def find_sheetname(
        self,
        path:          str,
        name_template: str
    ) -> str:
        workbook       = load_workbook(path)
        workbook_names = set(workbook.sheetnames)

        index = 0
        while name_template.format(index) in workbook_names:
            index += 1
        workbook.close()
        return name_template.format(index)

    def get_sheetnames(self, path: str) -> list[str]:
        workbook   = load_workbook(path)
        sheetnames = list(workbook.sheetnames)
        workbook.close()
        return sheetnames


writer_agent = WriterAgent()